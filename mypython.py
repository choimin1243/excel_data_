from openpyxl import load_workbook
import datetime

now = datetime.datetime.now()

now_after_7 = now + datetime.timedelta(days=7)

money_due_date = now_after_7.strftime('%Y-%m-%d')

read_wb = load_workbook('./test.xlsx', data_only=True)
read_ws = read_wb["Sheet1"]
read_ws_sheet2 = read_wb["Sheet2"]


print(read_ws_sheet2.cell(2,1).value)
print(read_ws_sheet2.cell(3,1).value)
print(read_ws_sheet2.cell(4,1).value)
print(read_ws_sheet2.cell(5,1).value)
print(read_ws_sheet2.cell(6,1).value)
print(read_ws_sheet2.cell(7,1).value)
print(read_ws_sheet2.cell(8,1).value)

samsung_email = read_ws_sheet2.cell(2,1).value
lg_email = read_ws_sheet2.cell(3,1).value
skt_email = read_ws_sheet2.cell(4,1).value
airbnb_email = read_ws_sheet2.cell(5,1).value
slack_email = read_ws_sheet2.cell(6,1).value
inflearn_email = read_ws_sheet2.cell(7,1).value
class101_email = read_ws_sheet2.cell(8,1).value

print(samsung_email)
print(lg_email)
print(skt_email)
print(airbnb_email)
print(slack_email)
print(inflearn_email)
print(class101_email)

list1 = []

row = 70

for i in range(2, row) :

    list1.append(read_ws.cell(i, 2).value)


list2 = set(list1)

list3 = list(filter(None, list2))


for i in list3 :

    read_wb_result = load_workbook("./result.xlsx")

    read_ws_result = read_wb_result["Sheet1"]

    current_company = i

    temp_list = []

    for j in range(2, row) :
        # j = 2,3,4,5,6,7 --

        if current_company == read_ws.cell(j, 2).value :
            temp_list.append(j)


    company = read_ws.cell(temp_list[0], 2).value
    owner = read_ws.cell(temp_list[0], 3).value
    user = read_ws.cell(temp_list[0], 4).value

    read_ws_result["B1"] = company
    read_ws_result["B2"] = owner
    read_ws_result["B3"] = user


    start_num = 12

    final_price = 0

    for oneOfTemp in temp_list :

        read_ws_result["A" + str(start_num)] = read_ws.cell(oneOfTemp, 1).value
        read_ws_result["B" + str(start_num)] = read_ws.cell(oneOfTemp, 2).value
        read_ws_result["C" + str(start_num)] = read_ws.cell(oneOfTemp, 4).value
        read_ws_result["D" + str(start_num)] = read_ws.cell(oneOfTemp, 7).value
        read_ws_result["E" + str(start_num)] = read_ws.cell(oneOfTemp, 8).value
        read_ws_result["F" + str(start_num)] = (read_ws.cell(oneOfTemp, 5).value +
                                                read_ws.cell(oneOfTemp, 6).value) * \
                                               (1 - read_ws.cell(oneOfTemp, 7).value) - \
                                               read_ws.cell(oneOfTemp, 8).value

        temp_price = (read_ws.cell(oneOfTemp, 5).value +read_ws.cell(oneOfTemp, 6).value) * (1 - read_ws.cell(oneOfTemp, 7).value) - read_ws.cell(oneOfTemp, 8).value

        if (read_ws.cell(oneOfTemp, 9).value == "Y") :
            temp_price = temp_price * 0.5

        read_ws_result["G" + str(start_num)] = temp_price

        start_num = start_num +1

        final_price = final_price + temp_price

    read_ws_result["B4"] = final_price

    read_wb_result.save("excel/" + company + ".xlsx")
    read_wb_result.close()

    email_price = format(int(final_price), ",")

    email_company = ""

    if company == "삼성" :
        email_company = samsung_email
    elif company == "엘지" :
        email_company = lg_email
    elif company == "SKT" :
        email_company = skt_email
    elif company == "Airbnb" :
        email_company = airbnb_email
    elif company == "슬랙" :
        email_company = slack_email
    elif company == "인프런" :
        email_company = inflearn_email
    elif company == "class101" :
        email_company = class101_email
    else:
        print("여기는 예외입니다!!")

    email_text = "이메일 : " + email_company + "\n" + \
                company + "회사 대표님 돈 주세용 \n\n" \
                + "사용 내역은 아래와 같습니다. \n\n" \
                + "담당자 : " + owner + "\n" \
                + "금액 : " + email_price + "원 \n" \
                + "사용자 : " + user + "\n\n" \
                + "오늘부터 7일 후인 " + money_due_date + "까지 입금해주세용"



    file = open('text/' + company + ".txt", mode="wt", encoding='utf-8')
    file.write(email_text)
    file.close()